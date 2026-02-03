# -*- coding: utf-8 -*-
"""Excel-related utility functions."""
from __future__ import annotations

import os
import re
from io import BytesIO
from typing import Any, Dict, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import NAME_KR_WIDTH, NAME_RU_WIDTH


def find_col_by_keyword(
    columns: Sequence[str],
    keyword: str,
    exclude: Optional[str] = None,
) -> Optional[str]:
    """Find first column containing keyword, optionally excluding pattern."""
    for col in columns:
        text = str(col)
        if keyword not in text:
            continue
        if exclude and exclude in text:
            continue
        return col
    return None


def clean_amount(value: Any) -> Any:
    """Normalize amount values to numeric or NA."""
    if pd.isna(value):
        return pd.NA
    text = str(value).strip()
    if not text:
        return pd.NA
    text = text.replace(" ", "")
    text = text.replace(",", ".")
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in {"", "-", ".", "-."}:
        return pd.NA
    return pd.to_numeric(text, errors="coerce")


def clean_amount_vectorized(series: pd.Series) -> pd.Series:
    """Vectorized amount normalization for a series."""
    if series.empty:
        return series
    mask = series.isna()
    text = series.astype(str).str.strip()
    text = text.str.replace(" ", "", regex=False)
    text = text.str.replace(",", ".", regex=False)
    text = text.str.replace(r"[^0-9.\-]", "", regex=True)
    text = text.replace({"": pd.NA, "-": pd.NA, ".": pd.NA, "-.": pd.NA})
    out = pd.to_numeric(text, errors="coerce")
    out = out.where(~mask, pd.NA)
    return out


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize column names by trimming and cleaning invisible characters."""
    cleaned = []
    for col in df.columns:
        text = str(col)
        text = text.replace("\ufeff", "")
        text = text.replace("\u200b", "")
        text = text.replace("\xa0", " ")
        text = text.strip()
        if "출결" in text and "출결여부" not in text:
            text = "출결여부"
        cleaned.append(text)
    df.columns = cleaned
    return df


def normalize_header_text(value: Any) -> str:
    """Normalize header text by removing whitespace and invisible chars."""
    text = str(value)
    text = text.replace("\ufeff", "")
    text = text.replace("\u200b", "")
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", "", text)
    return text


def read_excel_smart_bytes(file_bytes: bytes) -> pd.DataFrame:
    """Read Excel data with flexible header detection."""
    data = BytesIO(file_bytes)
    sheets = pd.read_excel(data, sheet_name=None, header=None)
    for sheet_name, sheet_df in sheets.items():
        max_scan = min(len(sheet_df), 20)
        for row_idx in range(max_scan):
            row = sheet_df.iloc[row_idx].tolist()
            normalized = [normalize_header_text(x) for x in row]
            if "고유번호" in normalized and "지역" in normalized:
                header = sheet_df.iloc[row_idx].tolist()
                df = sheet_df.iloc[row_idx + 1 :].copy()
                df.columns = header
                df = normalize_columns(df)
                df["__sheet"] = sheet_name
                return df

            if row_idx + 1 < max_scan:
                next_row = sheet_df.iloc[row_idx + 1].tolist()
                combined = [
                    normalize_header_text(a) + normalize_header_text(b)
                    for a, b in zip(row, next_row)
                ]
                if "고유번호" in combined and "지역" in combined:
                    df = sheet_df.iloc[row_idx + 2 :].copy()
                    df.columns = combined
                    df = normalize_columns(df)
                    df["__sheet"] = sheet_name
                    return df
    data.seek(0)
    df = pd.read_excel(data)
    df = normalize_columns(df)
    df["__sheet"] = 0
    return df


def read_excel_smart_path(file_path: str) -> pd.DataFrame:
    """Read Excel data from a file path with flexible header detection."""
    with open(file_path, "rb") as file:
        return read_excel_smart_bytes(file.read())


def list_excel_files(folder_path: str) -> List[str]:
    """List Excel files in a folder, excluding temporary files."""
    if not os.path.isdir(folder_path):
        return []
    files: List[str] = []
    for name in os.listdir(folder_path):
        if name.startswith("~$"):
            continue
        lower = name.lower()
        if lower.endswith(".xlsx") or lower.endswith(".xls"):
            files.append(os.path.join(folder_path, name))
    return sorted(files)


def list_yyyymm_subfolders(folder_path: str) -> List[str]:
    """List YYYY.MM subfolders under a path."""
    if not os.path.isdir(folder_path):
        return []
    results: List[str] = []
    for entry in os.scandir(folder_path):
        if not entry.is_dir():
            continue
        name = entry.name.strip()
        if re.match(r"^\d{4}\.\d{2}$", name):
            results.append(name)
    return sorted(results)


def make_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure dataframe columns are unique to avoid concat reindex errors."""
    seen: Dict[str, int] = {}
    new_cols: List[str] = []
    for col in df.columns:
        base = str(col)
        if base in seen:
            seen[base] += 1
            new_cols.append(f"{base}.{seen[base]}")
        else:
            seen[base] = 0
            new_cols.append(base)
    df.columns = new_cols
    return df


def _apply_header_style(ws: Any) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    header_font = Font(size=10, bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font


def _auto_size_columns(ws: Any) -> None:
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def _apply_name_widths(ws: Any) -> None:
    for col_idx, cell in enumerate(ws[1], start=1):
        header_text = normalize_header_text(cell.value).lower()
        if header_text == normalize_header_text("이름(kr)").lower():
            ws.column_dimensions[get_column_letter(col_idx)].width = NAME_KR_WIDTH
        elif header_text == normalize_header_text("이름(ru)").lower():
            ws.column_dimensions[get_column_letter(col_idx)].width = NAME_RU_WIDTH


def _apply_filter(ws: Any, df: pd.DataFrame, autofilter: Dict[str, Any]) -> None:
    max_row = ws.max_row
    max_col = ws.max_column
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    
    # "all": True 이거나 특정 컬럼 지정이 없으면 필터 화살표만 생성
    if autofilter.get("all"):
        return
        
    if "column" in autofilter and "value" in autofilter:
        col_name = autofilter["column"]
        if col_name in df.columns:
            filter_col = df.columns.get_loc(col_name)
            filter_value = autofilter.get("value")
            values = filter_value if isinstance(filter_value, list) else [filter_value]
            # openpyxl 0-based index for add_filter_column? No, it usually takes 0-based col id relative to range
            ws.auto_filter.add_filter_column(filter_col, values)


def _apply_invalid_uid_highlight(ws: Any, df: pd.DataFrame) -> None:
    """고유번호 형식이 00000000-00000 이 아닌 행을 빨강색으로 강조."""
    # 고유번호 컬럼 찾기
    uid_col_idx = -1
    for idx, col in enumerate(df.columns):
        if "고유번호" in str(col):
            uid_col_idx = idx
            break
            
    if uid_col_idx == -1:
        return

    # 스타일 정의 (진한 빨강 텍스트, 연한 빨강 배경)
    red_font = Font(color="9C0006", bold=True, size=10)
    pink_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    
    pattern = re.compile(r"^\d{8}-\d{5}$")
    
    # 데이터 행 순회 (헤더 제외, 2행부터)
    # df.iloc[row_idx]에 해당하는 엑셀 행은 row_idx + 2
    for i in range(len(df)):
        val = str(df.iloc[i, uid_col_idx]).strip()
        # 빈 값은 패스? 아니면 빈 값도 오류로 칠 것인가? -> 규칙4: "고유번호가 ... 아닌 경우"
        # 빈 값도 포함.
        is_valid = bool(pattern.match(val))
        
        if not is_valid:
            # 해당 행 전체 스타일 적용
            excel_row = i + 2
            # ws.iter_rows를 쓰기엔 비효율적일 수 있으므로 직접 접근
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=excel_row, column=col)
                cell.font = red_font
                cell.fill = pink_fill


def _apply_hide_rows(ws: Any, hide_rows: pd.Series) -> None:
    for row_idx, should_hide in enumerate(hide_rows, start=2):
        if bool(should_hide):
            ws.row_dimensions[row_idx].hidden = True


def apply_sheet_style(
    ws: Any,
    df: pd.DataFrame,
    autofilter: Optional[Dict[str, Any]] = None,
    hide_rows: Optional[pd.Series] = None,
    highlight_invalid_uid: bool = False,
) -> None:
    """Apply shared Excel styling to a worksheet."""
    font = Font(size=10)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
    _apply_header_style(ws)
    _auto_size_columns(ws)
    _apply_name_widths(ws)
    ws.freeze_panes = "A2"
    if autofilter:
        _apply_filter(ws, df, autofilter)
    if hide_rows is not None:
        _apply_hide_rows(ws, hide_rows)
    if highlight_invalid_uid:
        _apply_invalid_uid_highlight(ws, df)


def to_excel_bytes(
    df: pd.DataFrame,
    sheet_name: str = "Sheet1",
    autofilter: Optional[Dict[str, Any]] = None,
    hide_rows: Optional[pd.Series] = None,
    highlight_invalid_uid: bool = False,
) -> bytes:
    """Serialize dataframe to styled Excel bytes."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]
        apply_sheet_style(ws, df, autofilter=autofilter, hide_rows=hide_rows, highlight_invalid_uid=highlight_invalid_uid)
    return output.getvalue()


def write_excel_sheets(file_path: str, sheets: Sequence[Tuple[str, pd.DataFrame]]) -> None:
    """Write multiple sheets to an Excel file with shared styling."""
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.book[sheet_name]
            apply_sheet_style(ws, df)


def to_excel_multi_bytes(sheets: Sequence[Tuple[str, pd.DataFrame]]) -> bytes:
    """Serialize multiple sheets to styled Excel bytes."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.book[sheet_name]
            apply_sheet_style(ws, df)
    return output.getvalue()


def _report_column_widths(include_dept: bool) -> List[int]:
    """순번6, 언어권6, [부서6], 이름18, 고유번호18, 회비10, 체육10, 미납10, 십일조10, 미납10."""
    base = [6, 6]
    if include_dept:
        base.append(6)
    base.extend([18, 18, 10, 10, 10, 10, 10])
    return base


def _report_number_col_indices(include_dept: bool) -> Tuple[int, ...]:
    """1-based column indices for 회비, 체육회비 (오른쪽 정렬, 콤마)."""
    return (6, 7) if include_dept else (5, 6)


def to_report_excel_bytes(
    title: str,
    headers: List[str],
    data_rows: List[List[Any]],
    sheet_name: str = "Sheet1",
    include_dept: bool = False,
) -> bytes:
    """Report Excel: row 1 title (16pt, center, bold), row 2 empty, row 3 headers, row 4+ data.
    열너비·가운데/오른쪽 정렬·회비체육 숫자·테두리 적용."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ncols = max(len(headers), 1)
    if data_rows:
        ncols = max(ncols, len(data_rows[0]) if data_rows[0] else 0)
    end_col = get_column_letter(ncols)
    widths = _report_column_widths(include_dept)
    name_col_idx = 4 if include_dept else 3
    max_name_len = 0
    label_rows = {"장년회", "청년회"}
    for row in data_rows:
        if not row or row[0] in label_rows or len(row) < name_col_idx:
            continue
        val = row[name_col_idx - 1]
        if val is None:
            continue
        max_name_len = max(max_name_len, len(str(val)))
    if max_name_len:
        widths[name_col_idx - 1] = min(max(max_name_len + 2, widths[name_col_idx - 1]), 60)
    num_cols = _report_number_col_indices(include_dept)
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    ws["A1"] = title
    ws.merge_cells(f"A1:{end_col}1")
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = center
    for c in range(1, ncols + 1):
        ws.cell(row=2, column=c, value=None)

    header_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(size=10, bold=True)
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    label_rows = {"장년회", "청년회"}
    label_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    for r, row in enumerate(data_rows, start=4):
        is_label = bool(row) and row[0] in label_rows
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v if c == 1 or not is_label else None)
            if is_label:
                cell.font = Font(size=14, bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.fill = label_fill
                left_side = thin if c == 1 else None
                right_side = thin if c == ncols else None
                cell.border = Border(left=left_side, right=right_side, top=thin, bottom=thin)
            else:
                cell.font = Font(size=10)
                cell.alignment = right if c in num_cols else center
                if c in num_cols and v is not None:
                    cell.number_format = "#,##0"
                cell.border = border

    for col_idx, w in enumerate(widths, start=1):
        if col_idx <= ncols:
            ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A4"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
